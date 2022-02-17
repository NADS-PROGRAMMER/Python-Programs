<<<<<<< HEAD

# Creating patterns
sequenceIter = 1


for i in range(5):
    pattern = ''
    for j in range(sequenceIter):
        pattern += ' * '
    print(pattern)
    sequenceIter = sequenceIter + 1


# Iteration on list
grades = [100, 100, 100, 90, 94, 98]


for index in range(0, len(grades)):
    print(grades[index])
=======
import numpy as np
import openpyxl
import matplotlib.pyplot as plt

file = openpyxl.load_workbook('inventory.xlsx')

sheet = file['Sheet1']

companies = {}
inventory_per_company = {}

for i in range(2, sheet.max_row + 1):
    company = sheet.cell(i, 4).value
    inventory = sheet.cell(i, 3).value

    if company not in companies:
        companies[company] = company

array_companies = list(companies.values()) # convert to list



>>>>>>> f32eb80 (SOLVED Complements problem)
