import openpyxl as excel


class StudentGrades:
    def __init__(self, file_path):
        self._file_path = file_path

    def average_per_student(self):

        file = excel.load_workbook(self._file_path)

        grades = file['Grades']

        for row in range(2, grades.max_row + 1):

            student_name = grades.cell(row, 1).value
            student_average = (grades.cell(row, 2).value + grades.cell(row, 3).value + grades.cell(row, 4).value) / 3
            is_passed = 'Pass' if student_average >= 60 else 'Failed'

            result = f'{student_name} average is {student_average.__floor__()} = {is_passed}'

            print(result)

    '''
        <summary>
            This function determines who has the
            highest grade in the file.
        </summary> 
    '''
    def highest_grade(self):

        # load the file
        file = excel.load_workbook(self._file_path)

        # read the workbook sheet
        grades = file['Grades']

        # Initial values for highest_grade and student_name
        highest_grade = (grades.cell(2, 2).value + grades.cell(2, 3).value + grades.cell(2, 4).value) / 3
        student_name = grades.cell(2, 1).value

        for row in range(3, grades.max_row + 1):

            current_grade = (grades.cell(row, 2).value + grades.cell(row, 3).value + grades.cell(row, 4).value) / 3

            if current_grade > highest_grade:
                highest_grade = current_grade
                student_name = grades.cell(row, 1).value

        print(f'With Highest Grade: {student_name} {highest_grade.__floor__()}')

    def lowest_grade(self):

        # load the file
        file = excel.load_workbook(self._file_path)

        # read the workbook sheet
        grades = file['Grades']

        # Initial values for lowest_grade and student_name
        lowest_grade = (grades.cell(2, 2).value + grades.cell(2, 3).value + grades.cell(2, 4).value) / 3
        student_name = grades.cell(2, 1).value

        for row in range(3, grades.max_row + 1):

            current_grade = (grades.cell(row, 2).value + grades.cell(row, 3).value + grades.cell(row, 4).value) / 3

            if current_grade < lowest_grade:
                lowest_grade = current_grade
                student_name = grades.cell(row, 1).value

        print(f'With Lowest Grade: {student_name} {lowest_grade.__floor__()}')